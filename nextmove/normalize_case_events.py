from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from pykrx import stock


DATASET_ID = "nextmove_master"
TABLE_ID = "case_events"
PROJECT_ID = "infin-stock-bot"
DATE_LINE_RE = re.compile(r"^(?P<raw_date>\d{6})(?:\s+(?P<rest>.*))?$")
BODY_DATE_RE = re.compile(r"\((\d{2})\.(\d{2})\.(\d{2})\)")


@dataclass
class ParsedArticle:
    event_date: str
    market_tags: str | None
    article_title: str
    source_order: int


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    default_input = script_dir.parent / "종목_히스토리.xlsx"

    parser = argparse.ArgumentParser(
        description="Normalize 종목_히스토리.xlsx into BigQuery case_events rows."
    )
    parser.add_argument(
        "--input",
        default=str(default_input),
        help="Input xlsx path. Default: ../종목_히스토리.xlsx",
    )
    parser.add_argument(
        "--project",
        default=PROJECT_ID,
        help=f"GCP project ID. Default: {PROJECT_ID}",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse the workbook and export CSV without uploading to BigQuery.",
    )
    return parser.parse_args()


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def clean_theme(raw: object) -> str:
    text = normalize_text(raw) or ""
    text = re.sub(r"^\d+\.\s*", "", text)
    text = re.sub(r"\s*-\s*.*$", "", text)
    return re.sub(r"[\d\-]+$", "", text).strip()


def split_theme_lines(raw: object) -> list[str]:
    text = normalize_text(raw)
    if not text:
        return []

    themes: list[str] = []
    for line in str(text).replace("\r", "\n").split("\n"):
        cleaned = clean_theme(line)
        if cleaned:
            themes.append(cleaned)
    return themes


def normalize_core_theme(raw: object) -> str | None:
    themes = split_theme_lines(raw)
    if not themes:
        return None
    return themes[0]


def normalize_all_themes(raw: object) -> str | None:
    themes = split_theme_lines(raw)
    if not themes:
        return None
    seen: list[str] = []
    for theme in themes:
        if theme not in seen:
            seen.append(theme)
    return ",".join(seen)


def looks_like_theme_tags(text: str) -> bool:
    if not text or " " in text:
        return False
    parts = [part.strip() for part in text.split(",") if part.strip()]
    if not parts:
        return False
    allowed = re.compile(r"^[0-9A-Za-z가-힣+/&]+$")
    return all(allowed.match(part) and len(part) <= 12 for part in parts)


def normalize_event_date(raw_date: str) -> str:
    return datetime.strptime(raw_date, "%y%m%d").strftime("%Y-%m-%d")


def parse_article_line(line: str) -> tuple[str | None, str]:
    if " - " in line:
        head, title = line.split(" - ", 1)
        return (head.strip() or None), title.strip()
    if looks_like_theme_tags(line):
        return line.strip(), ""
    return None, line.strip()


def parse_article_block(raw: object) -> list[ParsedArticle]:
    text = normalize_text(raw)
    if not text:
        return []

    articles: list[ParsedArticle] = []
    current: ParsedArticle | None = None

    for order, raw_line in enumerate(text.replace("\r", "\n").split("\n")):
        line = raw_line.strip()
        if not line:
            continue

        matched = DATE_LINE_RE.match(line)
        if matched:
            event_date = normalize_event_date(matched.group("raw_date"))
            rest = (matched.group("rest") or "").strip()
            market_tags, article_title = parse_article_line(rest)
            current = ParsedArticle(
                event_date=event_date,
                market_tags=market_tags,
                article_title=article_title,
                source_order=order,
            )
            articles.append(current)
            continue

        if current is None:
            continue

        current.article_title = (
            f"{current.article_title}\n{line}".strip() if current.article_title else line
        )

    return articles


def parse_body_sections(raw: object) -> tuple[dict[str, str], str | None]:
    text = normalize_text(raw)
    if not text:
        return {}, None

    matches = list(BODY_DATE_RE.finditer(text))
    if not matches:
        return {}, text

    sections: dict[str, list[str]] = defaultdict(list)
    leading = text[: matches[0].start()].strip() or None

    for index, match in enumerate(matches):
        start = match.start()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        raw_date = "".join(match.groups())
        event_date = normalize_event_date(raw_date)
        section = text[start:end].strip()
        if section:
            sections[event_date].append(section)

    merged = {event_date: "\n\n".join(parts) for event_date, parts in sections.items()}
    return merged, leading


def determine_market_cond_from_tags(market_tags: str | None) -> str | None:
    tags = market_tags or ""
    if any(keyword in tags for keyword in ("급등", "상승")):
        return "상승"
    if any(keyword in tags for keyword in ("급락", "하락")):
        return "하락"
    return None

def fetch_market_cond_map(event_dates: set[str]) -> dict[str, str | None]:
    if not event_dates:
        return {}

    start = min(event_dates).replace("-", "")
    end = max(event_dates).replace("-", "")
    resolved: dict[str, str | None] = {event_date: None for event_date in event_dates}

    try:
        frame = stock.get_index_ohlcv_by_date(start, end, "1001")
        if frame.empty:
            return resolved
    except Exception:
        return resolved

    for index, series in frame.iterrows():
        event_date = index.strftime("%Y-%m-%d")
        if event_date not in resolved:
            continue

        if "등락률" in frame.columns:
            change_rate = float(series["등락률"])
        elif "종가" in frame.columns and "시가" in frame.columns:
            open_price = float(series["시가"])
            close_price = float(series["종가"])
            if open_price == 0:
                continue
            change_rate = ((close_price - open_price) / open_price) * 100
        else:
            continue

        if change_rate >= 0.5:
            resolved[event_date] = "상승"
        elif change_rate <= -0.5:
            resolved[event_date] = "하락"
        else:
            resolved[event_date] = "횡보"

    return resolved


def parse_workbook(input_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[dict[str, object]] = []
    unresolved_market_dates: set[str] = set()

    for excel_row in sheet.iter_rows(min_row=2, values_only=True):
        stock_name = normalize_text(excel_row[1] if len(excel_row) > 1 else None)
        article_raw = excel_row[3] if len(excel_row) > 3 else None
        if not stock_name or not normalize_text(article_raw):
            continue

        parsed_articles = parse_article_block(article_raw)
        if not parsed_articles:
            continue

        body_map, leading_body = parse_body_sections(excel_row[8] if len(excel_row) > 8 else None)
        core_theme = normalize_core_theme(excel_row[4] if len(excel_row) > 4 else None)
        all_themes = normalize_all_themes(excel_row[5] if len(excel_row) > 5 else None)
        is_leader = bool(normalize_text(excel_row[6] if len(excel_row) > 6 else None))
        keyword_summary = normalize_text(excel_row[7] if len(excel_row) > 7 else None)
        master_memo = normalize_text(excel_row[10] if len(excel_row) > 10 else None)

        for index, article in enumerate(parsed_articles):
            article_body = body_map.get(article.event_date)
            if article_body is None and index == 0 and leading_body:
                article_body = leading_body
            if article_body is None and len(parsed_articles) == 1 and leading_body:
                article_body = leading_body

            market_cond = determine_market_cond_from_tags(article.market_tags)
            if market_cond is None:
                unresolved_market_dates.add(article.event_date)

            rows.append(
                {
                    "case_id": f"{stock_name}_{article.event_date.replace('-', '')}",
                    "event_date": article.event_date,
                    "stock_code": None,
                    "stock_name": stock_name,
                    "core_theme": core_theme,
                    "all_themes": all_themes,
                    "article_title": article.article_title or None,
                    "article_body": article_body,
                    "market_cond": market_cond,
                    "is_leader": is_leader,
                    "rise_rate": None,
                    "trade_amount": None,
                    "keyword_summary": keyword_summary,
                    "master_memo": master_memo,
                }
            )

    market_cache = fetch_market_cond_map(unresolved_market_dates)
    for row in rows:
        if row["market_cond"] is None:
            row["market_cond"] = market_cache.get(str(row["event_date"]))

    workbook.close()
    return consolidate_rows(rows)


def merge_multiline_text(current: object, incoming: object, separator: str) -> str | None:
    parts: list[str] = []
    for value in (current, incoming):
        text = normalize_text(value)
        if text and text not in parts:
            parts.append(text)
    if not parts:
        return None
    return separator.join(parts)


def consolidate_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    merged: dict[str, dict[str, object]] = {}
    for row in rows:
        case_id = str(row["case_id"])
        existing = merged.get(case_id)
        if existing is None:
            merged[case_id] = dict(row)
            continue

        existing["article_title"] = merge_multiline_text(
            existing.get("article_title"),
            row.get("article_title"),
            "\n",
        )
        existing["article_body"] = merge_multiline_text(
            existing.get("article_body"),
            row.get("article_body"),
            "\n\n",
        )
        existing["keyword_summary"] = existing.get("keyword_summary") or row.get("keyword_summary")
        existing["master_memo"] = existing.get("master_memo") or row.get("master_memo")
        existing["core_theme"] = existing.get("core_theme") or row.get("core_theme")
        existing["all_themes"] = existing.get("all_themes") or row.get("all_themes")
        existing["market_cond"] = existing.get("market_cond") or row.get("market_cond")
        existing["is_leader"] = bool(existing.get("is_leader")) or bool(row.get("is_leader"))

    return list(merged.values())


def ensure_dataset_and_table(client: Any) -> Any:
    from google.cloud import bigquery
    from google.cloud.exceptions import NotFound

    dataset_ref = bigquery.DatasetReference(client.project, DATASET_ID)
    table_ref = dataset_ref.table(TABLE_ID)

    try:
        client.get_dataset(dataset_ref)
    except NotFound:
        dataset = bigquery.Dataset(dataset_ref)
        dataset.location = "US"
        client.create_dataset(dataset)

    schema = [
        bigquery.SchemaField("case_id", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("event_date", "DATE", mode="REQUIRED"),
        bigquery.SchemaField("stock_code", "STRING"),
        bigquery.SchemaField("stock_name", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("core_theme", "STRING"),
        bigquery.SchemaField("all_themes", "STRING"),
        bigquery.SchemaField("article_title", "STRING"),
        bigquery.SchemaField("article_body", "STRING"),
        bigquery.SchemaField("market_cond", "STRING"),
        bigquery.SchemaField("is_leader", "BOOL"),
        bigquery.SchemaField("rise_rate", "FLOAT"),
        bigquery.SchemaField("trade_amount", "FLOAT"),
        bigquery.SchemaField("keyword_summary", "STRING"),
        bigquery.SchemaField("master_memo", "STRING"),
    ]

    try:
        return client.get_table(table_ref)
    except NotFound:
        table = bigquery.Table(table_ref, schema=schema)
        return client.create_table(table)


def fetch_existing_case_ids(client: Any) -> set[str]:
    query = f"SELECT case_id FROM `{client.project}.{DATASET_ID}.{TABLE_ID}`"
    return {row["case_id"] for row in client.query(query).result()}


def export_csv(rows: Iterable[dict[str, object]], output_path: Path) -> None:
    fieldnames = [
        "case_id",
        "event_date",
        "stock_code",
        "stock_name",
        "core_theme",
        "all_themes",
        "article_title",
        "article_body",
        "market_cond",
        "is_leader",
        "rise_rate",
        "trade_amount",
        "keyword_summary",
        "master_memo",
    ]
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def upload_rows(client: Any, rows: list[dict[str, object]]) -> None:
    table_id = f"{client.project}.{DATASET_ID}.{TABLE_ID}"
    for start in range(0, len(rows), 500):
        chunk = rows[start : start + 500]
        errors = client.insert_rows_json(table_id, chunk)
        if errors:
            raise RuntimeError(f"BigQuery insert failed: {errors}")


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    parsed_rows = parse_workbook(input_path)
    if not parsed_rows:
        print("No case events were parsed.")
        return

    if args.dry_run:
        output_path = Path.cwd() / f"case_events_{datetime.now().strftime('%Y%m%d')}.csv"
        export_csv(parsed_rows, output_path)
        print(f"Dry-run complete: {len(parsed_rows)} rows exported to {output_path}")
        return

    from google.cloud import bigquery

    client = bigquery.Client(project=args.project)
    ensure_dataset_and_table(client)
    existing_case_ids = fetch_existing_case_ids(client)
    new_rows = [row for row in parsed_rows if row["case_id"] not in existing_case_ids]

    if not new_rows:
        print("No new rows to upload.")
        return

    upload_rows(client, new_rows)
    print(f"Upload complete: {len(new_rows)} new rows inserted.")


if __name__ == "__main__":
    main()
